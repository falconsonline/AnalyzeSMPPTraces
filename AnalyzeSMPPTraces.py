#!/usr/bin/env python3

import subprocess
import sys
import os
import re
import platform
import shutil
import tempfile
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Border, Side, Font

# ---------------- SMPP Dictionaries ----------------

TON_DESC={"0":"Unknown","1":"International","2":"National","3":"Network Specific","4":"Subscriber","5":"Alphanumeric","6":"Abbreviated"}
NPI_DESC={"0":"Unknown","1":"ISDN/E164","3":"Data","4":"Telex","6":"Land Mobile","8":"National","9":"Private"}
DATA_CODING_DESC={"0":"GSM Default","1":"ASCII","2":"Binary","3":"Latin-1","8":"UCS2"}
SMPP_STATUS_DESC={
"0x00000000":"Success",
"0x00000000":"No error (Success)",
"0x00000001":"Message length invalid",
"0x00000002":"Command length invalid",
"0x00000003":"Invalid command ID",
"0x00000004":"Incorrect bind state",
"0x00000005":"Already bound",
"0x00000008":"System error",
"0x0000000a":"Invalid source address",
"0x0000000b":"Invalid destination address",
"0x0000000d":"Bind failed",
"0x0000000e":"Invalid password",
"0x0000000f":"Invalid system ID",
"0x00000014":"Message queue full",
"0x00000033":"Invalid number of destinations",
"0x00000040":"Invalid destination flag",
"0x00000045":"Submit failed",
"0x00000058":"Throttling error (TPS exceeded)",
"0x000000c0":"Invalid optional parameter stream",
"0x000000c4":"Invalid optional parameter value",
"0x000000fe":"Delivery failure",
"0x000000ff":"Unknown error",
"0x0000110b":"Reserved",
"0x00000455":"Vendor-specific error"
}

# ---------------- Helpers ----------------

def normalize(v):
    if not v: return ""
    v=v.lower()
    if v.startswith("0x"): return str(int(v,16))
    return v

def desc(m,v): return m.get(normalize(v),"Unknown")

def status_desc(code):
    if not code: return ""
    code=code.lower()
    if code.startswith("0x"):
        try:
            dec=str(int(code,16))
            if dec in SMPP_STATUS_DESC:
                return SMPP_STATUS_DESC[dec]
        except: pass
    return SMPP_STATUS_DESC.get(code,"Unknown")

def excel_safe(v):
    if v is None: return ""
    v=str(v)
    v=ILLEGAL_CHARACTERS_RE.sub("",v)
    return v[:32760]

def is_hex(s):
    return bool(s and re.fullmatch(r'[0-9a-fA-F]+',s))

def decode_sms(msg,coding):
    if not msg: return ""
    if is_hex(msg):
        try:
            data=bytes.fromhex(msg)
            if coding in ["0x08","8"]:
                return data.decode("utf-16-be","ignore")
            return data.decode("latin1","ignore")
        except:
            return msg
    return msg

def format_time(epoch):
    return datetime.fromtimestamp(epoch).strftime("%Y-%m-%d %H:%M:%S.%f")[:-4]

def bucket_sort_key(b):
    if b.startswith("<"): return 0
    if b.startswith(">"): return 999
    if "-" in b:
        try: return float(b.split("-")[0])
        except: return 999
    return 999

def get_tool(name):
    if platform.system()=="Windows":
        default=f"C:\\Program Files\\Wireshark\\{name}.exe"
        if os.path.exists(default):
            return default
    return shutil.which(name) or name

tshark_path=get_tool("tshark")
mergecap_path=get_tool("mergecap")

# ---------------- INPUT ----------------

if len(sys.argv)<3:
    print("Usage: script.py <pcap|dir> output.xlsx [ports]")
    sys.exit(1)

input_path=sys.argv[1]
xls=sys.argv[2]
smpp_ports=sys.argv[3].split(",") if len(sys.argv)>=4 else []

# ---------------- PCAP LIST ----------------

pcap_files=[]
if os.path.isdir(input_path):
    for f in sorted(os.listdir(input_path)):
        if f.endswith(".pcap"):
            pcap_files.append(os.path.join(input_path,f))
else:
    pcap_files.append(input_path)

print("Processing PCAPs:",pcap_files)

# ---------------- OUTPUT DIR ----------------

timestamp=datetime.now().strftime("%Y%m%d_%H%M%S")
pcap_dir=f"failed_cases_pcaps/{timestamp}"
os.makedirs(pcap_dir,exist_ok=True)

# ---------------- DATA ----------------

submit_map=defaultdict(list)
responses=[]

submit_tps_conn=defaultdict(lambda: defaultdict(int))
total_tps_conn=defaultdict(lambda: defaultdict(int))
latency_ip_buckets=defaultdict(int)

msg_to_case={}
case_to_pcap={}

# ---------------- PROCESS ----------------

for pcap in pcap_files:

    decode_args=[]
    for p in smpp_ports:
        decode_args.extend(["-d",f"tcp.port=={p},smpp"])

    cmd=[tshark_path]
    cmd.extend(decode_args)

    cmd.extend([
        "-r",pcap,"-Y","smpp",
        "-T","fields","-E","separator=|",
        "-e","frame.time_epoch","-e","frame.number",
        "-e","ip.src","-e","ip.dst",
        "-e","tcp.srcport","-e","tcp.dstport",
        "-e","smpp.command_id","-e","smpp.sequence_number",
        "-e","smpp.service_type","-e","smpp.source_addr",
        "-e","smpp.source_addr_ton","-e","smpp.source_addr_npi",
        "-e","smpp.destination_addr","-e","smpp.dest_addr_ton",
        "-e","smpp.dest_addr_npi","-e","smpp.protocol_id",
        "-e","smpp.data_coding","-e","smpp.sm_length",
        "-e","smpp.message_text","-e","smpp.message",
        "-e","smpp.command_status","-e","smpp.sar_msg_ref_num",
        "-e","smpp.sar_segment_seqnum","-e","smpp.sar_total_segments"
    ])

    proc=subprocess.run(cmd,stdout=subprocess.PIPE,stderr=subprocess.PIPE,encoding="utf-8",errors="ignore")

    for line in proc.stdout.splitlines():

        p=line.split("|")
        if len(p)<21: continue

        ts,frame,ip_src,ip_dst,sport,dport,cmdid,seq,svc,src,ston,snpi,dst,dton,dnpi,pid,coding,sm_len,msg_text,msg,status,sar_ref,sar_seq,sar_total=p

        ts=float(ts)
        sec=int(ts)

        conn=f"{ip_src}:{sport}->{ip_dst}:{dport}"
        total_tps_conn[conn][sec]+=1
        if cmdid=="0x00000004":
            submit_tps_conn[conn][sec]+=1

        message = msg_text if msg_text else ""
        hex_decoded = decode_sms(msg,coding) if msg else ""

        key=(seq,ip_src,sport,ip_dst,dport)
        resp_key=(seq,ip_dst,dport,ip_src,sport)

        if cmdid=="0x00000004":
            submit_map[key].append({
                "time":ts,"frame":frame,"service":svc,
                "sender":src,"sender_ton":ston,"sender_npi":snpi,
                "recipient":dst,"rec_ton":dton,"rec_npi":dnpi,
                "pid":pid,"coding":coding,"sm_len":sm_len,
                "message":excel_safe(message),
                "hex_decoded":excel_safe(hex_decoded),
                "src_ip":ip_src,"dst_ip":ip_dst,
                "pcap":pcap,"sar_ref":sar_ref,"sar_seq":sar_seq,
                "sar_total":sar_total
            })

        elif cmdid=="0x80000004":
            responses.append({
                "key":resp_key,
                "time":ts,"frame":frame,
                "status":status,
                "pcap":pcap
            })

# ---------------- MATCH ----------------

rows=[]
group_times=defaultdict(list)

for r in responses:

    if r["key"] not in submit_map:
        continue

    best=None
    for s in submit_map[r["key"]]:
        if s["time"]<r["time"]:
            if best is None or s["time"]>best["time"]:
                best=s

    if not best:
        continue

    latency_sec = (r["time"]-best["time"])

    if latency_sec < 0.1: bucket="<0.1"
    elif latency_sec < 0.2: bucket="0.1-0.2"
    elif latency_sec < 0.3: bucket="0.2-0.3"
    elif latency_sec < 0.5: bucket="0.3-0.5"
    elif latency_sec < 1: bucket="0.5-1"
    elif latency_sec < 2: bucket="1-2"
    elif latency_sec < 3: bucket="2-3"
    elif latency_sec < 5: bucket="3-5"
    elif latency_sec < 10: bucket="5-10"
    else: bucket=">10"

    latency_ip_buckets[(bucket,best["src_ip"],best["dst_ip"])] += 1

    latency_ms = round(latency_sec*1000,2)

    msg = best["message"] if best["message"] else best["hex_decoded"]
    
    if best["sar_ref"]:
        g=f"SAR|{best['sar_ref']}"
    else:
        msg_val = best["message"] if best["message"] else best["hex_decoded"]
        g=f"SINGLE|{msg_val}"

# ---------------- TOTAL TIME ----------------
    group_total = {}

    for g, pairs in group_times.items():
    
        submit_times = [p[0] for p in pairs]
        resp_times   = [p[1] for p in pairs]
        total = max(resp_times) - min(submit_times)

        group_total[g] = round(total * 1000, 2)


    best["group"]=g
    group_times[g].append((best["time"], r["time"]))

    rows.append({
        "msg":msg,
        "submit":best,
        "resp":r,
        "latency":latency_ms
    })

# ---------------- FAILURE + PCAP ----------------

msg_success=defaultdict(list)
msg_fail=defaultdict(list)

for r in rows:
    if r["resp"]["status"]=="0x00000000":
        msg_success[r["msg"]].append(r)
    else:
        msg_fail[r["msg"]].append(r)

failure_rows=[]
case_id=1

for msg in msg_fail:

    fails=msg_fail[msg]
    succ=msg_success.get(msg,[])

    msisdn=fails[0]["submit"]["recipient"]
    safe_msisdn=re.sub(r'\W+','',msisdn)
    resp_code=fails[0]["resp"]["status"].replace("0x","")

    final_pcap=f"failed_case_{case_id}_{safe_msisdn}_{resp_code}.pcap"
    final_path=os.path.join(pcap_dir,final_pcap)

    msg_to_case[msg]=case_id
    case_to_pcap[case_id]=final_pcap

    failure_rows.append([
        case_id,msg,len(fails),len(succ),
        fails[0]["submit"]["frame"],fails[0]["resp"]["frame"],
        succ[0]["submit"]["frame"] if succ else "",
        succ[0]["resp"]["frame"] if succ else "",
        os.path.basename(fails[0]["submit"]["pcap"]),
        os.path.basename(fails[0]["resp"]["pcap"]),
        os.path.basename(succ[0]["submit"]["pcap"]) if succ else "",
        os.path.basename(succ[0]["resp"]["pcap"]) if succ else "",
        final_pcap,
        fails[0]["submit"]["src_ip"],
        fails[0]["submit"]["dst_ip"],
        "SMSC / routing transient issue" if succ else "Content filtering"
    ])

    pcap_frame_map=defaultdict(set)

    for f in fails:
        pcap_frame_map[f["submit"]["pcap"]].add(f["submit"]["frame"])
        pcap_frame_map[f["submit"]["pcap"]].add(f["resp"]["frame"])

    for s in succ[:3]:
        pcap_frame_map[s["submit"]["pcap"]].add(s["submit"]["frame"])
        pcap_frame_map[s["submit"]["pcap"]].add(s["resp"]["frame"])

    temp_files=[]

    for src_pcap,frames in pcap_frame_map.items():
        filt=" or ".join([f"frame.number=={f}" for f in sorted(frames,key=int)])

        tmp=tempfile.NamedTemporaryFile(delete=False,suffix=".pcap")
        tmp.close()

        subprocess.run([tshark_path,"-r",src_pcap,"-Y",filt,"-w",tmp.name])

        if os.path.exists(tmp.name) and os.path.getsize(tmp.name)>0:
            temp_files.append(tmp.name)

    if temp_files:
        subprocess.run([mergecap_path,"-w",final_path] + temp_files)
        for f in temp_files:
            os.unlink(f)

    case_id+=1

# ---------------- EXCEL ----------------

wb=Workbook()

ws=wb.active
ws.title="SMPP_Transactions"

headers=[
"Service Type","Sender","Sender TON","Sender TON Desc",
"Sender NPI","Sender NPI Desc","Recipient",
"Recipient TON","Recipient TON Desc","Recipient NPI","Recipient NPI Desc",
"Sequence#","Response Code","Response Desc","Protocol Id",
"Data Coding","Data Coding Desc","Message","Message(Hex Decoded)",
"SMPP SM Length","submit frame","resp frame",
"submit time","resp time","Latency ms","TOTAL_SUBMIT_TIME_MS",
"SAR_REF","SAR_SEQ","SAR_TOTAL",
"SUBMIT_PCAP","RESP_PCAP","FAILED_CASE_PCAP"
]

ws.append([h.upper() for h in headers])

for r in rows:
    p=r["submit"]

    case=msg_to_case.get(r["msg"],"")
    failed_pcap=case_to_pcap.get(case,"") if case else ""

    ws.append([
        p["service"],p["sender"],
        p["sender_ton"],desc(TON_DESC,p["sender_ton"]),
        p["sender_npi"],desc(NPI_DESC,p["sender_npi"]),
        p["recipient"],
        p["rec_ton"],desc(TON_DESC,p["rec_ton"]),
        p["rec_npi"],desc(NPI_DESC,p["rec_npi"]),
        r["resp"]["key"][0],
        r["resp"]["status"],status_desc(r["resp"]["status"]),
        p["pid"],
        p["coding"],desc(DATA_CODING_DESC,p["coding"]),
        p["message"],p["hex_decoded"],
        p["sm_len"],
        p["frame"],r["resp"]["frame"],
        format_time(p["time"]),format_time(r["resp"]["time"]),
        r["latency"],group_total.get(p["group"],0),
        p["sar_ref"],p["sar_seq"],p["sar_total"],
        os.path.basename(p["pcap"]),
        os.path.basename(r["resp"]["pcap"]),
        failed_pcap
    ])

ws2=wb.create_sheet("Failure_Analysis")
ws2.append([
"CASE_ID","MESSAGE","FAIL COUNT","SUCCESS COUNT",
"FAIL SUBMIT FRAME","FAIL RESP FRAME",
"SUCCESS SUBMIT FRAME","SUCCESS RESP FRAME",
"FAIL_SUBMIT_PCAP","FAIL_RESP_PCAP",
"SUCCESS_SUBMIT_PCAP","SUCCESS_RESP_PCAP",
"MERGED_FAILED_PCAP",
"SRC_IP","DST_IP","RCA"
])

for r in failure_rows:
    ws2.append(r)

ws3=wb.create_sheet("Latency_Distribution")
ws3.append(["BUCKET","COUNT","SRC_IP","DST_IP"])

grouped=defaultdict(list)
for (b,src,dst),v in latency_ip_buckets.items():
    grouped[(src,dst)].append((b,v))

for (src,dst) in sorted(grouped.keys()):
    for b,v in sorted(grouped[(src,dst)], key=lambda x: bucket_sort_key(x[0])):
        ws3.append([b,v,src,dst])

ws4=wb.create_sheet("Throughput_Analysis")
ws4.append(["TIME","CONNECTION","SUBMIT TPS","TOTAL TPS"])

for conn in sorted(total_tps_conn.keys()):
    for sec in sorted(total_tps_conn[conn].keys()):
        ws4.append([
            format_time(sec),
            conn,
            submit_tps_conn[conn].get(sec,0),
            total_tps_conn[conn][sec]
        ])

thin=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            cell.border=thin
    for cell in sheet[1]:
        cell.font=Font(bold=True)

wb.save(xls)

print("Report generated:",xls)
print("PCAP folder:",pcap_dir)
